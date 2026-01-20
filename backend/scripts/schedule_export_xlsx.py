import json
import sys
from datetime import datetime, time

from openpyxl import load_workbook


def excel_time_fraction(hms: str) -> str:
    s = (hms or "").strip()
    if not s:
        return ""
    try:
        parts = s.split(":")
        hh = int(parts[0])
        mm = int(parts[1]) if len(parts) > 1 else 0
        ss = int(parts[2]) if len(parts) > 2 else 0
        total = hh * 3600 + mm * 60 + ss
        frac = total / 86400.0
        # keep enough precision for Excel time formatting
        return f"{frac:.12f}".rstrip("0").rstrip(".")
    except Exception:
        return ""


def fmt_day_label(date_ymd: str, day_number: int) -> str:
    # Template style: YYYY.M.DD + newline + (N일차)
    try:
        d = datetime.strptime(date_ymd[:10], "%Y-%m-%d")
        return f"{d.year}.{d.month}.{d.day}\n({int(day_number)}일차)"
    except Exception:
        return f"{date_ymd}\n({int(day_number)}일차)"


def parse_hms(hms: str) -> time | None:
    s = (hms or "").strip()
    if not s:
        return None
    try:
        parts = s.split(":")
        hh = int(parts[0])
        mm = int(parts[1]) if len(parts) > 1 else 0
        ss = int(parts[2]) if len(parts) > 2 else 0
        return time(hour=hh, minute=mm, second=ss)
    except Exception:
        return None


def main() -> int:
    if len(sys.argv) < 4:
        print("usage: schedule_export_xlsx.py <template.xlsx> <input.json> <output.xlsx>", file=sys.stderr)
        return 2

    template_path = sys.argv[1]
    input_json = sys.argv[2]
    output_path = sys.argv[3]

    with open(input_json, "r", encoding="utf-8") as f:
        payload = json.load(f)

    # Use openpyxl to preserve template styles and Office namespaces.
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]

    # --- Top section mapping (based on template) ---
    ws["D7"].value = payload.get("productName", "") or ""
    ws["D8"].value = payload.get("tripPeriod", "") or ""
    # NOTE: In the template, B10:L10 is a merged cell.
    # Only the top-left cell (B10) is writable via openpyxl.
    ws["B10"].value = payload.get("meetingGuide", "") or ""

    mt = parse_hms(payload.get("meetingTime", "") or "")
    ws["D11"].value = mt

    ws["D12"].value = payload.get("meetingPlaceName", "") or ""
    ws["D13"].value = payload.get("meetingPlaceAddress", "") or ""

    # --- Per-day blocks (template has 3 blocks) ---
    day_blocks = [
        {  # day 1 block rows
            "day": 1,
            "date": "B18",
            "summary": "C18",
            "rows": [18, 19, 20],  # attractions slots
            "acc_name": "F23",
            "acc_addr": "K23",
            "transport": "F25",
            "meal_b": "K26",
            "meal_l": "M26",
            "meal_d": "O26",
        },
        {
            "day": 2,
            "date": "B29",
            "summary": "C29",
            "rows": [29, 30, 31],
            "acc_name": "F34",
            "acc_addr": "K34",
            "transport": "F36",
            "meal_b": "K37",
            "meal_l": "M37",
            "meal_d": "O37",
        },
        {
            "day": 3,
            "date": "B40",
            "summary": "C40",
            "rows": [40, 41, 42],
            "acc_name": "F45",
            "acc_addr": "K45",
            "transport": "F47",
            "meal_b": "K48",
            "meal_l": "M48",
            "meal_d": "O48",
        },
    ]

    days = payload.get("days") or []
    day_by_num = {int(d.get("day", 0)): d for d in days if int(d.get("day", 0)) > 0}

    for blk in day_blocks:
        dnum = int(blk["day"])
        day = day_by_num.get(dnum)
        if not day:
            ws[blk["date"]].value = None
            ws[blk["summary"]].value = None
            for r in blk["rows"]:
                ws[f"F{r}"].value = None
                ws[f"G{r}"].value = None
                ws[f"H{r}"].value = ""
                ws[f"K{r}"].value = ""
            ws[blk["acc_name"]].value = ""
            ws[blk["acc_addr"]].value = ""
            ws[blk["transport"]].value = ""
            ws[blk["meal_b"]].value = ""
            ws[blk["meal_l"]].value = ""
            ws[blk["meal_d"]].value = ""
            continue

        date_ymd = str(day.get("date") or "")
        ws[blk["date"]].value = fmt_day_label(date_ymd, dnum)
        ws[blk["summary"]].value = str(day.get("summary") or "")

        atts = list(day.get("attractions") or [])
        for i, r in enumerate(blk["rows"]):
            ref_start = f"F{r}"
            ref_end = f"G{r}"
            ref_name = f"H{r}"
            ref_addr = f"K{r}"
            if i < len(atts):
                a = atts[i] or {}
                ws[ref_start].value = parse_hms(str(a.get("start_time") or ""))
                ws[ref_end].value = parse_hms(str(a.get("end_time") or ""))
                ws[ref_name].value = str(a.get("name") or "")
                ws[ref_addr].value = str(a.get("address") or "")
            else:
                ws[ref_start].value = None
                ws[ref_end].value = None
                ws[ref_name].value = ""
                ws[ref_addr].value = ""

        ws[blk["acc_name"]].value = str(day.get("accommodation_name") or "")
        ws[blk["acc_addr"]].value = str(day.get("accommodation_address") or "")
        ws[blk["transport"]].value = str(day.get("transportation") or "")
        ws[blk["meal_b"]].value = str(day.get("breakfast") or "")
        ws[blk["meal_l"]].value = str(day.get("lunch") or "")
        ws[blk["meal_d"]].value = str(day.get("dinner") or "")

    wb.save(output_path)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())


